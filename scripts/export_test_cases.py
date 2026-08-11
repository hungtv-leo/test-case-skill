"""Xuat Excel ban giao cho tester: 2 sheet Verified + Gap.

Gate:
  - CHI block khi case coverage=verified chua pass.
  - Gap / exploratory / needs-product-decision KHONG block ban giao
    (dung de tester biet case code chua check).

Usage:
    python export_test_cases.py \
        --results results.json \
        --cases <path>/cases.json \
        --out .selftest_tmp/handover_<feature>.xlsx

Exit codes:
    0  -> xuat thanh cong (moi verified deu pass)
    2  -> verified con fail/error -> KHONG xuat
    3  -> loi input
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from _lib.outcomes import (  # noqa: E402
    category_label,
    coverage_label,
    get_coverage,
    outcomes_from_any,
    partition_cases,
    verified_gate_summary,
)
from _lib.validate import (  # noqa: E402
    ensure_cases,
    load_json,
    validate_outcomes_alignment,
)

VERIFIED_COLUMNS = [
    ("case_id", "Mã case", 14),
    ("description", "Mô tả", 40),
    ("precondition", "Điều kiện tiên đề", 30),
    ("steps", "Các bước", 45),
    ("data", "Dữ liệu", 30),
    ("expected", "Kết quả mong đợi", 35),
    ("actual", "Kết quả thực tế", 35),
    ("category", "Nhóm", 16),
    ("priority", "Ưu tiên", 10),
    ("status", "Trạng thái", 12),
]

GAP_COLUMNS = [
    ("case_id", "Mã case", 14),
    ("description", "Mô tả", 40),
    ("precondition", "Điều kiện tiên đề", 30),
    ("steps", "Các bước", 45),
    ("data", "Dữ liệu", 30),
    ("expected", "Kết quả mong đợi", 35),
    ("coverage", "Phân loại", 18),
    ("category", "Nhóm", 16),
    ("priority", "Ưu tiên", 10),
    ("code_evidence", "Bằng chứng code", 40),
    ("tester_note", "Ghi chú tester", 40),
    ("risk", "Rủi ro nếu bỏ sót", 35),
    ("status", "Trạng thái", 18),
]


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "\n".join(f"{i + 1}. {step}" for i, step in enumerate(value))
    if isinstance(value, dict):
        return "\n".join(f"{k}: {v}" for k, v in value.items())
    return str(value)


def _write_sheet(ws, columns: list, rows: list, header_color: str) -> None:
    header_fill = PatternFill("solid", fgColor=header_color)
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(vertical="top", wrap_text=True)

    for col_idx, (_, header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_normalize_cell(row.get(key)))
            cell.alignment = wrap

    ws.freeze_panes = "A2"


def _write_excel(verified_rows: list, gap_rows: list, out_path: str) -> None:
    wb = Workbook()
    ws_v = wb.active
    ws_v.title = "Đã verify"
    _write_sheet(ws_v, VERIFIED_COLUMNS, verified_rows, "4472C4")

    # Excel cấm ký tự / * ? : \\ [ ] trong tên sheet — dùng gạch ngang.
    ws_g = wb.create_sheet("Gap - Rủi ro")
    _write_sheet(ws_g, GAP_COLUMNS, gap_rows, "C65911")

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _build_verified_row(test_id: str, meta: dict, outcome: str | None) -> dict:
    row = dict(meta)
    row["case_id"] = meta.get("case_id", test_id)
    row["coverage"] = coverage_label(get_coverage(meta))
    row["category"] = category_label(meta.get("category"))
    row["actual"] = meta.get("expected", "") if outcome == "passed" else f"[{outcome}]"
    row["status"] = "Đạt" if outcome == "passed" else "Không đạt"
    return row


def _build_gap_row(test_id: str, meta: dict, outcome: str | None) -> dict:
    row = dict(meta)
    row["case_id"] = meta.get("case_id", test_id)
    row["coverage"] = coverage_label(get_coverage(meta))
    row["category"] = category_label(meta.get("category"))
    if outcome is None:
        row["status"] = "Lỗ hổng (chưa tự động hóa)"
    elif outcome == "passed":
        row["status"] = "Đã cover"
    else:
        row["status"] = f"Lỗ hổng ({outcome})"
    return row


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Xuat Excel ban giao (Verified + Gap). Gate: chi verified phai pass."
    )
    parser.add_argument("--results", required=True, help="results.json chuan hoac pytest .report.json")
    parser.add_argument("--cases", required=True, help="metadata cases.json (khoa theo test id)")
    parser.add_argument("--out", required=True, help="duong dan file .xlsx dau ra")
    args = parser.parse_args()

    try:
        raw_results = load_json(args.results)
        raw_cases = load_json(args.cases)
    except (FileNotFoundError, ValueError) as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        sys.exit(3)

    cases = ensure_cases(raw_cases, args.cases)
    try:
        outcomes = outcomes_from_any(raw_results)
    except ValueError as exc:
        print(f"[ERROR] {args.results}: {exc}", file=sys.stderr)
        sys.exit(3)

    alignment = validate_outcomes_alignment(cases, outcomes)
    if alignment:
        print("[ERROR] cases.json va results khong khop (verified bat buoc co ket qua):", file=sys.stderr)
        for line in alignment:
            print(line, file=sys.stderr)
        sys.exit(3)

    all_verified_ok, passed, total, problems, _ids, gap_count = verified_gate_summary(
        outcomes, cases
    )
    if not all_verified_ok:
        print("[FAIL] Con case VERIFIED chua pass -> KHONG xuat Excel. Bao dev sua:", file=sys.stderr)
        for case_id, test_id, outcome in problems:
            print(f"  - {case_id} ({test_id}): {outcome}", file=sys.stderr)
        sys.exit(2)

    verified, gaps = partition_cases(cases)
    verified_rows = [
        _build_verified_row(tid, meta, outcomes.get(tid)) for tid, meta in verified.items()
    ]
    gap_rows = [_build_gap_row(tid, meta, outcomes.get(tid)) for tid, meta in gaps.items()]

    _write_excel(verified_rows, gap_rows, args.out)
    print(
        f"[OK] Verified {passed}/{total} pass. Gap/rủi ro: {gap_count}. Da xuat: {args.out}"
    )


if __name__ == "__main__":
    main()
